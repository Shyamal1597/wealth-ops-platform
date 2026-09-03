import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\SSFL-RETAIL-017\Downloads\Sunidhi_Accessibility_Remediation_Status.xlsx', data_only=False)
ws = wb['About - About Story']
header_row = None
for r in range(1, 8):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v and str(v).strip() == 'Test Result' for v in vals):
        header_row = r
        break
headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(header_row, c).value
    if v: headers[str(v).strip()] = c
print("HEADERS:", headers)
def get(row, colname):
    c = headers.get(colname)
    if not c: return None
    cell = ws.cell(row, c)
    return {'value': cell.value, 'hyperlink': cell.hyperlink.target if cell.hyperlink else None}
CLEAN = {'pass','dna','n/a','na','not applicable',''}
for r in range(header_row+1, ws.max_row+1):
    tr = get(r, 'Test Result')
    tr_val = tr['value'] if tr else None
    tr_norm = str(tr_val).strip().lower() if tr_val is not None else None
    sc = get(r, 'Success Criteria'); err = get(r, 'Error Description')
    def real(field):
        if not field or field['value'] is None: return False
        s = str(field['value']).replace('\xa0','').strip()
        return len(s) >= 2
    flagged = False
    if tr_norm is None:
        flagged = real(sc) or real(err)
    elif tr_norm not in CLEAN:
        flagged = True
    if flagged:
        print(r, get(r,'Sr. #'), sc, tr, get(r,'Impact/Severity'), get(r,'Screenshot'), err, get(r,'Remediation Status'), get(r,'Fix Description / Root Cause'), get(r,'Verification Evidence'), sep=' | ')
